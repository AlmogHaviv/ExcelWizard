import pandas as pd
import os
import re
import numpy as np
from openpyxl import load_workbook
import shutil
from datetime import datetime
import argparse


def setup_arg_parser():
    """
    Set up the argument parser.
    """
    parser = argparse.ArgumentParser(description='parse for type of report')
    parser.add_argument('-all', '--all_of_the_reports', action='store_true', help='If entered, return the algo report')
    parser.add_argument('-product', '--product_report', action='store_true', help='If entered, return the algo report')
    parser.add_argument('-devops', '--devops_report', action='store_true', help='If entered, return the dev report')
    parser.add_argument('-algo', '--algo_report', action='store_true', help='If entered, return the algo report')
    parser.add_argument('-dev', '--dev_report', action='store_true', help='If entered, return the dev report')
    parser.add_argument('-bi', '--bi_report', action='store_true', help='If entered, return the bi report')
    parser.add_argument('-sa', '--system_architect_report', action='store_true', help='If entered, return the bi report')
    parser.add_argument('-cpb', '--CPBDirectors_report', action='store_true', help='If entered, return the bi report')
    parser.add_argument('-input_file', '--input_file', type=str, help='Input file')
    parser.add_argument('-wd', '--wd', type=str, help='Working directory')
    return parser


def main():
    try:
        """
        Main logic of the script using parsed arguments.
        """
        parser = setup_arg_parser()

        # Parse the arguments
        args = parser.parse_args()

        # Check which report type was selected and perform corresponding action
        if args.all_of_the_reports:
            print("Generating all of the report...")
            sub_main(args.wd, args.input_file, "all")
        elif args.bi_report:
            print("Generating the Bi report...")
            sub_main(args.wd, args.input_file, "Bi")
        elif args.dev_report:
            print("Generating the Dev report...")
            sub_main(args.wd, args.input_file, "Dev")
        elif args.algo_report:
            print("Generating the Algo report...")
            sub_main(args.wd, args.input_file, "Algo")
        elif args.product_report:
            print("Generating the Product report...")
            sub_main(args.wd, args.input_file, "Product")
        elif args.devops_report:
            print("Generating the Devops report...")
            sub_main(args.wd, args.input_file, "Devops")
        elif args.system_architect_report:
            print("Generating the SystemArchitect report...")
            sub_main(args.wd, args.input_file, "SystemArchitect")
        elif args.CPBDirectors_report:
            print("Generating the CPBDirectors report...")
            sub_main(args.wd, args.input_file, "CPBDirectors")
        else:
            print("please specify the type of report, use -h to know which types there are")

    except UnicodeDecodeError as e:
        print("An exception occurred:", str(e))


def sub_main(wd, filename, arg):
    input_path = os.path.join('..', f'{wd}', f'{filename}')

    # Creates a dict of algo, bi, dev, devops and product with relevant df for each one of them.
    dict_of_dfs = create_df_for_cpbc(input_path)

    for department, df in dict_of_dfs.items():
        if arg == department or arg == "all":
            print(f"{department} report is in process")
            create_full_scale_for_excel(department, df, wd)
            print(f"{department} report is ready")

    print("All the reports are ready")
    return


def create_df_for_cpbc(path):
    try:
        # Read the execl file to create a df to work with
        expanded_df = pd.read_excel(path)

        # Identify the Sprint columns
        sprint_columns = [col for col in expanded_df.columns if 'Sprint' in col]

        # Determine the latest non-null date for each row in the Sprint columns
        expanded_df['Latest Sprint Date'] = expanded_df[sprint_columns].max(axis=1)

        # fetching the necessary data from the expended df
        df = expanded_df[['Time Spent', 'Assignee', 'Latest Sprint Date', 'Custom field (Budget)']]
        df['Time Spent (Days)'] = df['Time Spent'] / 3600 / 8
        df.drop(columns=['Time Spent'], inplace=True)

        # rename to keep it same as the cpbc report
        df.rename(columns={'Latest Sprint Date': 'Sprint'}, inplace=True)

        # Format the date to 'YYYY-MM'
        df['Sprint'] = df['Sprint'].dt.strftime('%m-%Y')

        df_by_departments = departments_df(df)

        return df_by_departments

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None


def departments_df(df: object):
    # Construct the relative path to the CSV file
    csv_file_path = os.path.join('..', 'config', 'worker-names.csv')
    xlsx_file_path = os.path.join('..', 'config', 'fte_contract.csv')

    # Read the CSV file
    worker_config = pd.read_csv(csv_file_path)

    # Read the CSV file
    fte_contract = pd.read_csv(xlsx_file_path)

    # Create a mapping from assignee names to their corresponding columns in df1
    name_to_column = {}
    for column in worker_config.columns:
        for name in worker_config[column].dropna():
            name_to_column[name] = column

    # Add the corresponding column information to df2
    df['Team'] = df['Assignee'].map(name_to_column)

    # Drop rows where 'Time Spent' is NaN
    df = df.dropna(subset=['Time Spent (Days)'])

    # Initialize a dictionary to store the DataFrames
    category_dfs = {}

    # Loop through each category column in df1
    for category in worker_config.columns:
        # get planned per month
        fte_col = fte_contract[['month', category]]
        fte_col['month'] = pd.to_datetime(fte_col['month'])
        # Get the list of assignees for the current category
        assignees = worker_config[category].dropna().tolist()
        # Filter df2 based on the assignees list
        filtered_df = df[df['Assignee'].isin(assignees)]
        # Apply the clean_string function to the 'Custom field (Budget)' column
        filtered_df['Custom field (Budget)'] = filtered_df['Custom field (Budget)'].apply(clean_string)
        filtered_df.drop(columns=['Team'], inplace=True)
        # Group by 'Sprint' and 'Custom field (Budget)' and sum the 'Time Spent (Days)'
        df_aggregated = filtered_df.groupby(['Sprint', 'Custom field (Budget)'], as_index=False)['Time Spent (Days)'].sum()
        # Pivot the DataFrame without aggregation
        pivot_df = df_aggregated.pivot(index='Sprint', columns='Custom field (Budget)', values='Time Spent (Days)')
        # Fill NaN values with 0
        pivot_df = pivot_df.fillna(0)
        # Reset index to convert the pivoted DataFrame back to a normal DataFrame
        pivot_df = pivot_df.reset_index()
        # Calculate the sum across the specified columns for each row
        pivot_df['Total Time Spent'] = pivot_df.iloc[:, 2:].sum(axis=1)
        # Let's extract the year-month format from the 'Custom field (Budget)' column in df2
        pivot_df['month'] = pd.to_datetime(pivot_df['Sprint'])
        # Now, merge the two datasets on the 'month' column
        merged_df = pd.merge(pivot_df, fte_col, on='month', how='left')
        # Rename the 'Dev' column to 'fte_contract'
        merged_df.rename(columns={category: 'FTE Contract'}, inplace=True)
        merged_df['OH'] = merged_df['FTE Contract'] - merged_df['Total Time Spent']
        merged_df = merged_df.dropna(subset=['OH'])
        merged_df.drop(columns=['month'], inplace=True)
        merged_df.rename(columns={'Sprint': 'Month'}, inplace=True)
        category_dfs[category] = merged_df

    return category_dfs


def clean_string(input_string):
    if input_string[:4] == "P000":
        return "P0 - Vacation / Sickness"
    elif input_string[:6] == "P999 -":
        return "P999 - General2"
    elif "P999" in input_string and "(Biomica)" in input_string:
        return "P999 - General4"
    elif "P145-Corteva - IA" in input_string:
        return "P145 - Corteva"
    elif "P86 -Product" in input_string:
        return "P86 - Product"
    elif "P264 - Product-CP (Chempass)" == input_string:
        return "P264 -Product CP"
    elif "P84" in input_string:
        return "p84-New program"
    elif "P85 -Syngenta (Lavie)" == input_string:
        return "P85 - Syngenta"
    elif "P192 - LAV 321 (Lavie)" == input_string:
        return "P192 - LAV 321"
    elif "P274 - Product- Upkeep ChemPass" in input_string:
        return "P274 - Product- Upkeep CP "
    elif "P165 - VERB BIOTICS" == input_string:
        return "P165 - Verb Biotics"
    elif "P401 - The Kitchen" == input_string:
        return "P401 - The Kitchen"
    elif "P403 - Run Generator (on going) - Casterra " == input_string:
        return "P403 - Casterra RUN Generator"
    elif "P213 - Breeding general  (Canonic 2023)" == input_string:
        return "P213 - Breeding general "
    elif "P285 - Ag Plenus" == input_string:
        return "P285 - DevOps - CP"
    elif "P275 - Experimental Upkeep (CPB)" == input_string:
        return "P275 - CPB Upkeep Experimental"
    elif "P295 - GCP MIGRATION (CPB)" == input_string:
        return "P295 - Migration to GCP MB & (maybe GR)"
    else:
        # Use regex to remove the content inside parentheses along with the space before it
        cleaned_string = re.sub(r'\s*\(.*?\)', '', input_string)
    return cleaned_string


def create_full_scale_for_excel(department, df, wd):
    print(department)
    # Get the current date and time
    current_date = datetime.now()

    # Convert the date column to datetime
    df['Month'] = pd.to_datetime(df['Month'], format='%m-%Y')

    # Extract the current month and year
    current_month = current_date.month
    current_year = current_date.year
    if current_month == 1:
        current_year -= 1

    department_name = department
    df['Entry Type'] = 'Actual'
    df['Employee ID'] = None
    df['Exp Type'] = '951 - Ongoing Task'
    df['Jira name'] = None
    df['Employee Name'] = 'Total'
    df['Approved by'] = 'Ilia Zhidkov'
    df['FTE left to Assign'] = 0
    if department == "Bi":
        department_name = "Bioinformation"
        df['Department'] = "405 - Bioinformatics"
        df['Role Ending'] = "T103 - Bioinformatician"
    elif department == "Algo":
        department_name = "Algo"
        df['Department'] = "404 - Algorithm"
        df['Role Ending'] = "T102 - Algorithm Developer"
    elif department == "Dev":
        department_name = "SoftwareDevelopment"
        df['Department'] = "406 - Software Development"
        df['Role Ending'] = "T112 - Software Developer"
    elif department == "Devops":
        department_name = "DevOps"
        df['Department'] = "420 - DevOps"
        df['Role Ending'] = "T105 - DevOps"
    elif department == "SystemArchitect":
        department_name = "System architect"
        df['Department'] = "422 -CPB Directors"
        df['Role Ending'] = "T115 - System Architect"
        df['Exp Type'] = '950 - Development Task'
    elif department == "CPBDirectors":
        department_name = "CPBDirectors"
        df['Department'] = "422 -CPB Directors"
        df['Role Ending'] = "CPB Directors"
        df['Exp Type'] = '950 - Development Task'
        df['Approved by'] = "Mark Kapel"
    else:
        df['Department'] = f"{department}"
        df['Role Ending'] = f"{department}"

    # fetching the template to work with in order to create the final report
    template_path = os.path.join('..', 'config', 'template_for_CPBC.xlsx')

    template_df = pd.read_excel(template_path, skiprows=1)

    columns = template_df.columns.tolist()

    # Create a list to store the new rows
    new_rows = []

    # Iterate over each row of the original DataFrame
    for index, row in df.iterrows():
        # Initialize a dictionary to store values for each row
        row_values = {}

        # Iterate over each column in the new DataFrame
        for column in columns:
            # Check if the column exists in the original DataFrame
            if column in df.columns:
                # If the column exists, assign the corresponding value from the original DataFrame
                row_values[column] = row[column]
            else:
                # If the column doesn't exist, assign NaN
                row_values[column] = np.nan

        # Append the row values to the list of new rows
        new_rows.append(row_values)

    # Convert the list of new rows to a DataFrame and concatenate with the template DataFrame
    new_rows_df = pd.DataFrame(new_rows)
    template_df = pd.concat([template_df, new_rows_df], ignore_index=True)
    template_df['OH'] = template_df['OH'] - template_df['P0 - Vacation / Sickness']

    # Path to save the new Excel file
    new_file_path = os.path.join('..', f'{wd}', f'{department_name}_{current_month - 1}_{str(current_year)[2:]}.xlsx')

    # Copy the template to a new location
    shutil.copy(template_path, new_file_path)

    # Load the workbook and select the active sheet
    wb = load_workbook(new_file_path)
    ws = wb.active

    # Write the DataFrame to the Excel sheet starting from the third row
    for row_index, data_row in template_df.iterrows():
        for col_index, value in enumerate(data_row):
            # Check if the value is a timestamp, convert to datetime object
            if isinstance(value, pd.Timestamp):
                formatted_date = value.strftime('%Y-%m')
                ws.cell(row=row_index + 3, column=col_index + 1, value=formatted_date)
            else:
                ws.cell(row=row_index + 3, column=col_index + 1, value=value)

    if department == "Bi" and current_date.year == 2024:
        # Specify the row and column of the cell you want to modify
        row = 5
        column = 68
        new_value = 14.3

        # Modify the cell value
        ws.cell(row=row, column=column).value = new_value

    # Save the workbook
    wb.save(new_file_path)

    print(template_df)

    return


if __name__ == "__main__":
    main()
