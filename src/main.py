import pandas as pd
import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import utils
import datetime


def read_excel_file(file_path):
    """
    Read data from an Excel file, perform necessary transformations, and prepare the data for further processing.

    Args:
    - file_path (str): Path to the Excel file.

    Returns:
    - tuple: A tuple containing a list of DataFrames grouped by role, and the headers row DataFrame.
    """
    try:
        # Read the first row to retrieve headers information
        first = pd.read_excel(file_path)
        headers_row = first.iloc[[0]]
        headers_row.insert(0, 'Unnamed - 1', '')
        headers_row.insert(0, 'Unnamed - 2', '')

        # Read the Excel files, skipping the first row to avoid duplication
        df = pd.read_excel(file_path, header=1)

        # Drop unnecessary columns
        columns_to_drop = ['Department Units', 'M.W.D', 'month', 'Special notes']
        df.drop(columns=columns_to_drop, inplace=True)

        # Add new columns with default values
        df['Entry Type'] = 'ACTUAL'
        df['Employee ID'] = None
        df['Exp Type'] = 'Ongoing task'
        df['Jira name'] = None
        df['Employee Name'] = 'Total'
        df['Approved by'] = None
        df['New Project 1'] = None
        df['New Project 2'] = None

        # Reorganize the order of columns
        cols = list(df.columns)
        cols.insert(0, cols.pop(cols.index('Approved by')))
        cols.insert(0, cols.pop(cols.index('Exp Type')))
        cols.insert(0, cols.pop(cols.index('Employee Name')))
        cols.insert(0, cols.pop(cols.index('Jira name')))
        cols.insert(0, cols.pop(cols.index('Department')))
        cols.insert(0, cols.pop(cols.index('Role Ending')))
        cols.insert(0, cols.pop(cols.index('Exp Type')))
        cols.insert(0, cols.pop(cols.index('Employee ID')))
        cols.insert(0, cols.pop(cols.index('Entry Type')))
        df = df[cols]

        # Group the concatenated DataFrame by 'Department'
        grouped_data = df.groupby('Department')

        # Initialize a list to store the 6 DataFrames
        grouped_dfs = []

        # Iterate over the groups and create a DataFrame for each group
        for group_name, group_df in grouped_data:
            if group_name != "Total":
                # Append the DataFrame for the current group to the list
                grouped_dfs.append([shorten_name(group_name), group_df])

        return grouped_dfs, headers_row

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None


def manipulate_data(data, first_row, output_directory):
    """
    Perform data manipulation tasks.

    Args:
    - data (pandas DataFrame): DataFrame containing the input data.

    Returns:
    - pandas DataFrame: DataFrame containing the manipulated data.
    """
    # Get a list of the headers
    header_row = first_row.columns.tolist()
    for idx, name in enumerate(header_row):
        if "Unnamed" == name[:7]:
            header_row[idx] = None
    header_row2 = data[0][1].columns.tolist()
    # Iterate over each DataFrame
    for idx, df in enumerate(data):
        # Create a new Excel workbook
        wb = Workbook()

        # Create a new sheet
        ws = wb.active

        # Add the first row from my_row to the Excel sheet
        ws.append(header_row)

        # Add the first row from df to the Excel sheet
        ws.append(header_row2)

        # Append all other rows from df to the Excel sheet
        for row in df[1].itertuples(index=False):
            ws.append(list(row))

        style_excel(ws)

        # Get the current date and time
        current_date = datetime.datetime.now()

        # Extract the current month and year
        current_month = current_date.month
        current_year = current_date.year
        if current_month == 1:
            current_year -= 1

        # Define the path for the Excel file
        if df[0] == "PLM" or df[0] == "PlantGrowth":
            file_name = f"{output_directory}/{df[0]}_{current_month - 1}_{current_year}_U.xlsx"
        elif df[0] == "QA":
            file_name = f"{output_directory}/PLM_{current_month - 1}_{current_year} QA.xlsx"
        else:
            file_name = f"{output_directory}/{df[0]}_{current_month - 1}_{current_year}.xlsx"

        # Save the workbook to an Excel file
        wb.save(file_name)

        print(file_name)


# Define a function to extract the shortened name using regex
def shorten_name(name):
    split_string = re.split(r'\s+|-', name)
    return "".join(split_string[1:])


def style_excel(ws):
    # Define fill and font styles
    light_blue_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    soft_green_fill = PatternFill(start_color='D3EAD3', end_color='D3EAD3', fill_type='solid')
    david_font = Font(name='David')

    # Apply styles to the second row
    for cell in ws[2]:
        cell.fill = light_blue_fill

    # Apply styles to columns 2-5 from the third row to the end
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=11):
        for cell in row:
            cell.fill = soft_green_fill

    # Apply font style to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = david_font

    # Add black borders between all parts of the data
    max_row = ws.max_row
    max_col = ws.max_column
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            ws.cell(row=i, column=j).border = thin_border

    # Define color fills
    mid_blue_fill = PatternFill(start_color="1F89DF", end_color="1F89DF", fill_type="solid")
    light_red_fill = PatternFill(start_color="DF761F", end_color="DF761F", fill_type="solid")
    pale_orange_fill = PatternFill(start_color="F3C68A", end_color="F3C68A", fill_type="solid")
    pale_blue_fill = PatternFill(start_color="6EEAF3", end_color="6EEAF3", fill_type="solid")
    light_green_fill = PatternFill(start_color="6EF396", end_color="6EF396", fill_type="solid")
    pale_yellow_fill = PatternFill(start_color="E9F483", end_color="E9F483", fill_type="solid")
    marker_blue_fill = PatternFill(start_color="4AB3EE", end_color="4AB3EE", fill_type="solid")
    marker_green_fill = PatternFill(start_color="55FE33", end_color="55FE33", fill_type="solid")
    marker_pink_fill = PatternFill(start_color="FC5CE6", end_color="FC5CE6", fill_type="solid")

    # Define column ranges and their respective fills
    column_fills = {
        ('L', 'Q'): mid_blue_fill,
        ('R', 'Z'): light_red_fill,
        ('AA', 'AD'): pale_orange_fill,
        ('AE', 'AI'): pale_blue_fill,
        ('AJ', 'AO'): light_green_fill,
        ('AP', 'AY'): pale_yellow_fill,
        ('AZ', 'BE'): marker_blue_fill,
        ('BF', 'BF'): marker_green_fill,
        ('BG', 'BG'): marker_pink_fill
    }

    # Apply colors to the specified columns in the first row
    for (start_col, end_col), fill_color in column_fills.items():
        start_index = openpyxl.utils.column_index_from_string(start_col)
        end_index = openpyxl.utils.column_index_from_string(end_col)
        for col in range(start_index, end_index + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_color


def main():
    input_file_path = os.path.abspath("../data/Experimental_Project_days_template.xlsx")
    output_file_path = os.path.abspath("../data/")
    # Step 1: Read input Excel file
    input_data, first_row = read_excel_file(input_file_path)
    utils.main()

    if input_data is None:
        return

    # Step 2: excel it
    manipulate_data(input_data, first_row, output_file_path)


if __name__ == "__main__":
    main()
